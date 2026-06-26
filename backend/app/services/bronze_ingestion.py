from __future__ import annotations

import csv
from datetime import date, datetime
from decimal import Decimal
from hashlib import sha256
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any
from uuid import UUID

from openpyxl import load_workbook
from pypdf import PdfReader
from sqlalchemy.orm import Session

from app.core.config import Settings, get_settings
from app.repositories.bronze_repository import BronzeRepository
from app.schemas.ingestion import FileUploadResponse, ImportBatchDetail, RawFileDetail
from app.services.source_detection import (
    SUPPORTED_EXTENSIONS,
    SourceDetection,
    detect_sicoob_pdf_source_from_text,
    detect_source,
)


class BronzeIngestionError(Exception):
    pass


class UnsupportedFileTypeError(BronzeIngestionError):
    pass


class UploadTooLargeError(BronzeIngestionError):
    pass


SOURCE_TYPE_EXTENSIONS = {
    "mercado_livre_account_statement_csv": ".csv",
    "mercado_livre_manual_cdb_csv": ".csv",
    "manual_investment_csv": ".csv",
    "b3_monthly_consolidated_xlsx": ".xlsx",
    "b3_annual_consolidated_xlsx": ".xlsx",
    "sicoob_checking_statement_pdf": ".pdf",
    "sicoob_card_invoice_pdf": ".pdf",
    "sicoob_investments_pdf": ".pdf",
    "sicoob_payroll_pdf": ".pdf",
}

SOURCE_TYPE_INSTITUTIONS = {
    "mercado_livre_account_statement_csv": "mercado_livre",
    "mercado_livre_manual_cdb_csv": "mercado_livre",
    "b3_monthly_consolidated_xlsx": "b3",
    "b3_annual_consolidated_xlsx": "b3",
    "sicoob_checking_statement_pdf": "sicoob",
    "sicoob_card_invoice_pdf": "sicoob",
    "sicoob_investments_pdf": "sicoob",
    "sicoob_payroll_pdf": "sicoob",
}


def _safe_error_message(exc: Exception) -> str:
    return f"{exc.__class__.__name__}: raw extraction failed"


def _json_value(value: Any) -> Any:
    if isinstance(value, datetime | date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return value


class BronzeIngestionService:
    def __init__(self, session: Session, settings: Settings | None = None) -> None:
        self.session = session
        self.settings = settings or get_settings()
        self.repository = BronzeRepository(session)

    def ingest_upload(
        self,
        *,
        filename: str,
        content: bytes,
        mime_type: str | None,
        source_type_override: str | None = None,
    ) -> FileUploadResponse:
        if not content:
            raise BronzeIngestionError("Uploaded file is empty.")
        if len(content) > self.settings.max_upload_size_bytes:
            raise UploadTooLargeError(
                f"Uploaded file exceeds the configured limit of {self.settings.max_upload_size_bytes} bytes."
            )

        extension = Path(filename).suffix.lower()
        if extension not in SUPPORTED_EXTENSIONS:
            raise UnsupportedFileTypeError(f"Unsupported file extension: {extension or '<none>'}.")

        file_hash = sha256(content).hexdigest()
        detection = detect_source(filename, mime_type=mime_type)
        if extension == ".pdf" and detection.source_type == "sicoob_pdf_unknown":
            detection = self._detect_pdf_source_from_content(content)
        if source_type_override:
            detection = self._source_detection_override(
                source_type=source_type_override,
                extension=extension,
                fallback=detection,
            )
        existing_raw_file = self.repository.find_raw_file_by_hash(file_hash)

        try:
            if existing_raw_file is not None:
                batch = self.repository.create_import_batch(
                    raw_file_id=existing_raw_file["id"],
                    source_type=source_type_override or existing_raw_file["source_type"] or detection.source_type,
                    status="duplicate",
                    total_records=0,
                    valid_records=0,
                    invalid_records=0,
                )
                self.repository.add_file_metadata(
                    raw_file_id=existing_raw_file["id"],
                    key="duplicate_upload",
                    value={
                        "uploaded_filename": filename,
                        "sha256_hash": file_hash,
                        "import_batch_id": str(batch["id"]),
                        "source_type_override": source_type_override,
                    },
                )
                self.session.commit()
                return self._build_upload_response(
                    raw_file=existing_raw_file,
                    batch=batch,
                    duplicate=True,
                )

            stored_path = self._store_raw_file(content=content, file_hash=file_hash, extension=extension)
            raw_file = self.repository.create_raw_file(
                original_filename=filename,
                stored_path=str(stored_path),
                mime_type=mime_type,
                file_extension=extension,
                file_size_bytes=len(content),
                sha256_hash=file_hash,
                source_type=detection.source_type,
                detected_institution=detection.detected_institution,
                metadata={
                    "storage_kind": "local",
                    "hash_algorithm": "sha256",
                    "safe_filename": stored_path.name,
                    "source_type_override": source_type_override,
                },
            )
            self.repository.add_file_metadata(
                raw_file_id=raw_file["id"],
                key="source_detection",
                value={
                    "source_type": detection.source_type,
                    "detected_institution": detection.detected_institution,
                    "file_extension": extension,
                    "mime_type": mime_type,
                    "source_type_override": source_type_override,
                },
            )
            batch = self.repository.create_import_batch(
                raw_file_id=raw_file["id"],
                source_type=detection.source_type,
                status="upload_received",
            )

            try:
                total_records = self._extract_raw_content(
                    import_batch_id=batch["id"],
                    extension=extension,
                    content=content,
                )
            except Exception as exc:
                message = _safe_error_message(exc)
                batch = self.repository.finish_import_batch(
                    batch_id=batch["id"],
                    status="failed",
                    total_records=0,
                    valid_records=0,
                    invalid_records=1,
                    error_message=message,
                )
                self.repository.add_parser_error(
                    import_batch_id=batch["id"],
                    error_type="raw_extraction_error",
                    error_message=message,
                    raw_reference={"raw_file_id": str(raw_file["id"])},
                    payload={"file_extension": extension, "source_type": detection.source_type},
                )
                self.repository.update_raw_file_status(raw_file["id"], "failed")
                self.session.commit()
                raise BronzeIngestionError(message) from exc

            batch = self.repository.finish_import_batch(
                batch_id=batch["id"],
                status="raw_extracted",
                total_records=total_records,
                valid_records=total_records,
                invalid_records=0,
            )
            self.repository.update_raw_file_status(raw_file["id"], "raw_extracted")
            raw_file = self.repository.get_raw_file(raw_file["id"]) or raw_file
            self.session.commit()
            return self._build_upload_response(raw_file=raw_file, batch=batch, duplicate=False)
        except BronzeIngestionError:
            raise
        except Exception:
            self.session.rollback()
            raise

    def list_raw_files(self, limit: int) -> list[dict[str, Any]]:
        return self.repository.list_raw_files(limit=limit)

    def get_raw_file(self, raw_file_id: UUID) -> dict[str, Any] | None:
        return self.repository.get_raw_file(raw_file_id)

    def get_raw_file_detail(self, raw_file_id: UUID) -> RawFileDetail | None:
        raw_file = self.repository.get_raw_file(raw_file_id)
        if raw_file is None:
            return None
        batches = self.repository.list_import_batches_for_raw_file(raw_file_id)
        raw_counts = self.repository.count_raw_records(batches[0]["id"]) if batches else {}
        return RawFileDetail(**raw_file, import_batches=batches, raw_counts=raw_counts)

    def list_import_batches(self, limit: int) -> list[ImportBatchDetail]:
        details: list[ImportBatchDetail] = []
        for batch in self.repository.list_import_batches(limit=limit):
            detail = self.get_import_batch(batch["id"])
            if detail is not None:
                details.append(detail)
        return details

    def get_import_batch(self, batch_id: UUID) -> ImportBatchDetail | None:
        batch = self.repository.get_import_batch(batch_id)
        if batch is None:
            return None
        raw_file = self.repository.get_raw_file(batch["raw_file_id"])
        if raw_file is None:
            return None
        raw_counts = self.repository.count_raw_records(batch["id"])
        return ImportBatchDetail(**batch, raw_file=raw_file, raw_counts=raw_counts)

    def _store_raw_file(self, *, content: bytes, file_hash: str, extension: str) -> Path:
        storage_dir = Path(self.settings.file_storage_path) / "raw"
        storage_dir.mkdir(parents=True, exist_ok=True)
        stored_path = storage_dir / f"{file_hash}{extension}"
        if not stored_path.exists():
            stored_path.write_bytes(content)
        return stored_path

    def _extract_raw_content(self, *, import_batch_id: UUID, extension: str, content: bytes) -> int:
        if extension == ".csv":
            return self.repository.insert_csv_rows(
                import_batch_id=import_batch_id,
                rows=self._extract_csv_rows(content),
            )
        if extension == ".xlsx":
            return self.repository.insert_xlsx_rows(
                import_batch_id=import_batch_id,
                rows=self._extract_xlsx_rows(content),
            )
        if extension == ".pdf":
            return self.repository.insert_pdf_pages(
                import_batch_id=import_batch_id,
                pages=self._extract_pdf_pages(content),
            )
        raise UnsupportedFileTypeError(f"Unsupported file extension: {extension}.")

    def _source_detection_override(
        self,
        *,
        source_type: str,
        extension: str,
        fallback: SourceDetection,
    ) -> SourceDetection:
        expected_extension = SOURCE_TYPE_EXTENSIONS.get(source_type)
        if expected_extension is None:
            raise UnsupportedFileTypeError(f"Unsupported source type: {source_type}.")
        if expected_extension != extension:
            raise UnsupportedFileTypeError(
                f"Source type {source_type} expects {expected_extension}, got {extension}."
            )
        return SourceDetection(
            source_type=source_type,
            detected_institution=SOURCE_TYPE_INSTITUTIONS.get(source_type, fallback.detected_institution),
        )

    def _extract_csv_rows(self, content: bytes) -> list[dict[str, Any]]:
        text = self._decode_text(content)
        rows: list[dict[str, Any]] = []
        for row_number, raw_text in enumerate(text.splitlines(), start=1):
            parsed_row = next(csv.reader(StringIO(raw_text)), [])
            rows.append(
                {
                    "row_number": row_number,
                    "raw_text": raw_text,
                    "raw_payload": {"columns": parsed_row},
                }
            )
        return rows

    def _extract_xlsx_rows(self, content: bytes) -> list[dict[str, Any]]:
        workbook = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
        rows: list[dict[str, Any]] = []
        for worksheet in workbook.worksheets:
            for row_number, values in enumerate(worksheet.iter_rows(values_only=True), start=1):
                normalized_values = [_json_value(value) for value in values]
                if all(value in (None, "") for value in normalized_values):
                    continue
                raw_text = " | ".join("" if value is None else str(value) for value in normalized_values)
                rows.append(
                    {
                        "sheet_name": worksheet.title,
                        "row_number": row_number,
                        "raw_text": raw_text,
                        "raw_payload": {"values": normalized_values},
                    }
                )
        workbook.close()
        return rows

    def _extract_pdf_pages(self, content: bytes) -> list[dict[str, Any]]:
        reader = PdfReader(BytesIO(content))
        pages: list[dict[str, Any]] = []
        for index, page in enumerate(reader.pages, start=1):
            text = page.extract_text() or ""
            pages.append(
                {
                    "page_number": index,
                    "extracted_text": text,
                    "extraction_metadata": {"page_number": index},
                }
            )
        return pages

    def _detect_pdf_source_from_content(self, content: bytes) -> SourceDetection:
        try:
            reader = PdfReader(BytesIO(content))
            page_texts = []
            for index, page in enumerate(reader.pages):
                if index >= 4:
                    break
                page_texts.append(page.extract_text() or "")
            return detect_sicoob_pdf_source_from_text("\n".join(page_texts))
        except Exception:
            return SourceDetection(source_type="sicoob_pdf_unknown", detected_institution="sicoob")

    def _decode_text(self, content: bytes) -> str:
        for encoding in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                return content.decode(encoding)
            except UnicodeDecodeError:
                continue
        raise BronzeIngestionError("CSV file encoding is not supported.")

    def _build_upload_response(
        self,
        *,
        raw_file: dict[str, Any],
        batch: dict[str, Any],
        duplicate: bool,
    ) -> FileUploadResponse:
        raw_counts = self.repository.count_raw_records(batch["id"])
        return FileUploadResponse(
            raw_file=raw_file,
            import_batch=batch,
            duplicate=duplicate,
            raw_counts=raw_counts,
        )
