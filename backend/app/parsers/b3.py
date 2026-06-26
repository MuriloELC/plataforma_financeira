from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
import re
from typing import Any
from uuid import UUID

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

from app.parsers.base import ParsedDocument, ParserError, make_record
from app.parsers.utils import MONTHS, normalize_text, parse_date, parse_decimal


class _B3BaseParser:
    supported_extensions = {".xlsx"}

    def detect(self, file_path: Path, metadata: dict[str, Any] | None = None) -> bool:
        del metadata
        return file_path.suffix.lower() == ".xlsx"

    def _load_workbook(self, file_path: Path) -> Workbook:
        workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
        for worksheet in workbook.worksheets:
            if getattr(worksheet, "max_column", None) == 1 and hasattr(worksheet, "reset_dimensions"):
                worksheet.reset_dimensions()
        return workbook

    def _parse_positions(
        self,
        *,
        workbook: Workbook,
        import_batch_id: UUID,
        source_file_id: UUID | None,
    ) -> tuple[dict[str, Decimal], list[Any]]:
        totals: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
        records = []

        for worksheet in workbook.worksheets:
            asset_class = self._asset_class_for_sheet(worksheet.title)
            if asset_class is None:
                continue

            rows = list(worksheet.iter_rows(values_only=True))
            if not rows:
                continue
            header = [str(value) if value is not None else "" for value in rows[0]]
            product_index = self._header_index(
                header,
                ("produto",),
                code="missing_b3_product_column",
                message="B3 product column was not found.",
            )
            institution_index = self._optional_header_index(header, ("instituicao",))
            value_index = self._value_index(header)

            for row_number, row in enumerate(rows[1:], start=2):
                product = self._cell(row, product_index)
                if product in (None, ""):
                    continue
                amount = parse_decimal(self._cell(row, value_index))
                totals[asset_class] += amount
                data = {
                    "asset_class": asset_class,
                    "product": str(product),
                    "institution": (
                        str(self._cell(row, institution_index))
                        if institution_index is not None and self._cell(row, institution_index) is not None
                        else None
                    ),
                    "amount": amount,
                }
                records.append(
                    make_record(
                        source_type=self.source_type,
                        import_batch_id=import_batch_id,
                        source_file_id=source_file_id,
                        data=data,
                        raw_reference={"sheet_name": worksheet.title, "raw_row": row_number},
                    )
                )

        return dict(totals), records

    def _parse_income(
        self,
        *,
        workbook: Workbook,
        import_batch_id: UUID,
        source_file_id: UUID | None,
    ) -> tuple[Decimal, list[Any], list[date]]:
        worksheet = self._worksheet_by_name(workbook, "proventos recebidos")
        if worksheet is None:
            return Decimal("0"), [], []

        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            return Decimal("0"), [], []
        header = [str(value) if value is not None else "" for value in rows[0]]
        product_index = self._header_index(
            header,
            ("produto",),
            code="missing_b3_income_product_column",
            message="B3 income product column was not found.",
        )
        payment_index = self._header_index(
            header,
            ("pagamento", "data de pagamento"),
            code="missing_b3_income_payment_column",
            message="B3 income payment column was not found.",
        )
        income_type_index = self._header_index(
            header,
            ("tipo de evento", "tipo"),
            code="missing_b3_income_type_column",
            message="B3 income type column was not found.",
        )
        amount_index = self._header_index(
            header,
            ("valor liquido", "valor"),
            code="missing_b3_income_value_column",
            message="B3 income value column was not found.",
        )
        records = []
        dates = []
        total = Decimal("0")
        for row_number, row in enumerate(rows[1:], start=2):
            product = self._cell(row, product_index)
            if not row or product in (None, ""):
                continue
            payment_date = self._parse_cell_date(self._cell(row, payment_index))
            amount = parse_decimal(self._cell(row, amount_index))
            total += amount
            dates.append(payment_date)
            records.append(
                make_record(
                    source_type=self.source_type,
                    import_batch_id=import_batch_id,
                    source_file_id=source_file_id,
                    data={
                        "record_type": "income",
                        "product": str(product),
                        "payment_date": payment_date.isoformat(),
                        "income_type": str(self._cell(row, income_type_index)),
                        "net_amount": amount,
                    },
                    raw_reference={"sheet_name": worksheet.title, "raw_row": row_number},
                )
            )
        return total, records, dates

    def _parse_trades(
        self,
        *,
        workbook: Workbook,
        import_batch_id: UUID,
        source_file_id: UUID | None,
    ) -> tuple[int, list[Any], list[date]]:
        worksheet = self._worksheet_by_name(workbook, "negociacoes")
        if worksheet is None:
            return 0, [], []

        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            return 0, [], []
        header = [str(value) if value is not None else "" for value in rows[0]]
        ticker_index = self._header_index(
            header,
            ("codigo de negociacao",),
            code="missing_b3_trade_ticker_column",
            message="B3 trade ticker column was not found.",
        )
        date_index = self._header_index(
            header,
            ("periodo inicial", "data"),
            code="missing_b3_trade_date_column",
            message="B3 trade date column was not found.",
        )
        institution_index = self._header_index(
            header,
            ("instituicao",),
            code="missing_b3_trade_institution_column",
            message="B3 trade institution column was not found.",
        )
        buy_index = self._header_index(
            header,
            ("quantidade compra",),
            code="missing_b3_trade_buy_column",
            message="B3 trade buy quantity column was not found.",
        )
        sell_index = self._header_index(
            header,
            ("quantidade venda",),
            code="missing_b3_trade_sell_column",
            message="B3 trade sell quantity column was not found.",
        )
        net_index = self._header_index(
            header,
            ("quantidade liquida",),
            code="missing_b3_trade_net_column",
            message="B3 trade net quantity column was not found.",
        )
        records = []
        dates = []
        for row_number, row in enumerate(rows[1:], start=2):
            ticker = self._cell(row, ticker_index)
            if not row or ticker in (None, ""):
                continue
            trade_date = self._parse_cell_date(self._cell(row, date_index))
            dates.append(trade_date)
            records.append(
                make_record(
                    source_type=self.source_type,
                    import_batch_id=import_batch_id,
                    source_file_id=source_file_id,
                    data={
                        "record_type": "trade",
                        "ticker": str(ticker),
                        "trade_date": trade_date.isoformat(),
                        "institution": str(self._cell(row, institution_index)),
                        "buy_quantity": parse_decimal(self._cell(row, buy_index)),
                        "sell_quantity": parse_decimal(self._cell(row, sell_index)),
                        "net_quantity": parse_decimal(self._cell(row, net_index)),
                    },
                    raw_reference={"sheet_name": worksheet.title, "raw_row": row_number},
                )
            )
        return len(records), records, dates

    def _asset_class_for_sheet(self, sheet_name: str) -> str | None:
        normalized = normalize_text(sheet_name)
        if "renda fixa" in normalized:
            return "renda_fixa"
        if "etf" in normalized:
            return "etf"
        if "posicao" in normalized and "acoes" in normalized:
            return "acao"
        return None

    def _value_index(self, header: list[str]) -> int:
        return self._header_index(
            header,
            ("valor atualizado curva", "valor atualizado mtm", "valor atualizado"),
            code="missing_b3_value_column",
            message="B3 value column was not found.",
        )

    def _header_index(
        self,
        header: list[str],
        candidates: tuple[str, ...],
        *,
        code: str,
        message: str,
    ) -> int:
        normalized = [self._normalize_header(value) for value in header]
        for candidate in candidates:
            normalized_candidate = self._normalize_header(candidate)
            if normalized_candidate in normalized:
                return normalized.index(normalized_candidate)
        raise ParserError(code, message)

    def _optional_header_index(self, header: list[str], candidates: tuple[str, ...]) -> int | None:
        try:
            return self._header_index(
                header,
                candidates,
                code="missing_optional_b3_column",
                message="Optional B3 column was not found.",
            )
        except ParserError:
            return None

    def _normalize_header(self, value: str) -> str:
        return re.sub(r"[^a-z0-9]+", " ", normalize_text(value)).strip()

    def _cell(self, row: tuple[Any, ...], index: int | None) -> Any:
        if index is None or index >= len(row):
            return None
        return row[index]

    def _parse_cell_date(self, value: Any) -> date:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        return parse_date(str(value))

    def _worksheet_by_name(self, workbook: Workbook, *candidates: str) -> Any | None:
        normalized_candidates = {self._normalize_header(candidate) for candidate in candidates}
        for worksheet in workbook.worksheets:
            if self._normalize_header(worksheet.title) in normalized_candidates:
                return worksheet
        return None

    def _reference_date_from_filename(self, file_path: Path) -> date | None:
        normalized = normalize_text(file_path.stem)
        year_match = re.search(r"(20\d{2})", normalized)
        if year_match is None:
            return None
        year = int(year_match.group(1))
        for month_name, month in MONTHS.items():
            if re.search(rf"\b{re.escape(month_name)}\b", normalized):
                return date(year, month, 1)
        return None


class B3MonthlyConsolidatedXlsxParser(_B3BaseParser):
    source_type = "b3_monthly_consolidated_xlsx"

    def parse(
        self,
        file_path: Path,
        import_batch_id: UUID,
        source_file_id: UUID | None = None,
    ) -> ParsedDocument:
        workbook = self._load_workbook(file_path)
        try:
            positions_by_class, position_records = self._parse_positions(
                workbook=workbook,
                import_batch_id=import_batch_id,
                source_file_id=source_file_id,
            )
            income_total, income_records, income_dates = self._parse_income(
                workbook=workbook,
                import_batch_id=import_batch_id,
                source_file_id=source_file_id,
            )
            trades_count, trade_records, trade_dates = self._parse_trades(
                workbook=workbook,
                import_batch_id=import_batch_id,
                source_file_id=source_file_id,
            )
        finally:
            workbook.close()

        reference_date = (
            max(income_dates + trade_dates)
            if income_dates or trade_dates
            else self._reference_date_from_filename(file_path)
        )
        if reference_date is None:
            raise ParserError("missing_b3_reference_month", "B3 monthly report has no date reference.")

        return ParsedDocument(
            source_type=self.source_type,
            import_batch_id=import_batch_id,
            source_file_id=source_file_id,
            payload={
                "reference_month": f"{reference_date.year:04d}-{reference_date.month:02d}",
                "positions_by_class": positions_by_class,
                "income_received_total": income_total,
                "trades_count": trades_count,
                "rules": {
                    "b3_is_official_for_listed_assets": True,
                    "b3_fixed_income_counts_as_reserve": False,
                },
            },
            records=[*position_records, *income_records, *trade_records],
            raw_reference={"file_path": str(file_path)},
        )


class B3AnnualConsolidatedXlsxParser(_B3BaseParser):
    source_type = "b3_annual_consolidated_xlsx"

    def parse(
        self,
        file_path: Path,
        import_batch_id: UUID,
        source_file_id: UUID | None = None,
    ) -> ParsedDocument:
        workbook = self._load_workbook(file_path)
        try:
            positions_by_class, position_records = self._parse_positions(
                workbook=workbook,
                import_batch_id=import_batch_id,
                source_file_id=source_file_id,
            )
            created = workbook.properties.created
        finally:
            workbook.close()

        reference_year = str((created.year - 1) if created else date.today().year - 1)
        return ParsedDocument(
            source_type=self.source_type,
            import_batch_id=import_batch_id,
            source_file_id=source_file_id,
            payload={
                "reference_year": reference_year,
                "positions_by_class": positions_by_class,
                "rules": {"annual_report_is_snapshot": True},
            },
            records=position_records,
            raw_reference={"file_path": str(file_path)},
        )
