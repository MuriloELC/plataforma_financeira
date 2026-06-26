import json
import os
from decimal import Decimal
from pathlib import Path
import re
from uuid import UUID
from zipfile import ZipFile

import pytest
from openpyxl import Workbook

from app.parsers import (
    B3AnnualConsolidatedXlsxParser,
    B3MonthlyConsolidatedXlsxParser,
    MercadoLivreAccountStatementCsvParser,
    MercadoLivreManualCdbCsvParser,
    ParserError,
    SicoobCardInvoicePdfParser,
    SicoobCheckingStatementPdfParser,
    SicoobInvestmentsPdfParser,
    SicoobPayrollPdfParser,
)

IMPORT_BATCH_ID = UUID("00000000-0000-0000-0000-000000000001")
SOURCE_FILE_ID = UUID("00000000-0000-0000-0000-000000000002")


def fixture_path(relative_path: str) -> Path:
    return Path(os.getenv("FIXTURES_PATH", "fixtures")) / relative_path


def expected_payload(relative_path: str) -> dict:
    with fixture_path(relative_path).open("r", encoding="utf-8") as file_handle:
        return json.load(file_handle)


@pytest.mark.parametrize(
    ("parser", "fixture", "expected"),
    (
        (
            MercadoLivreAccountStatementCsvParser(),
            "anonymized/mercado_livre/account_statement_sample.csv",
            "expected/mercado_livre_account_statement_expected.json",
        ),
        (
            MercadoLivreManualCdbCsvParser(),
            "anonymized/mercado_livre/cdb_position_sample.csv",
            "expected/cdb_position_expected.json",
        ),
        (
            B3MonthlyConsolidatedXlsxParser(),
            "anonymized/b3/relatorio-consolidado-mensal-sample.xlsx",
            "expected/b3_monthly_expected.json",
        ),
        (
            B3AnnualConsolidatedXlsxParser(),
            "anonymized/b3/relatorio-consolidado-anual-sample.xlsx",
            "expected/b3_annual_expected.json",
        ),
        (
            SicoobPayrollPdfParser(),
            "anonymized/sicoob/contracheque_sample.pdf",
            "expected/sicoob_payroll_expected.json",
        ),
        (
            SicoobCheckingStatementPdfParser(),
            "anonymized/sicoob/extrato_conta_sample.pdf",
            "expected/sicoob_checking_statement_expected.json",
        ),
        (
            SicoobCardInvoicePdfParser(),
            "anonymized/sicoob/fatura_cartao_sample.pdf",
            "expected/sicoob_card_invoice_expected.json",
        ),
        (
            SicoobInvestmentsPdfParser(),
            "anonymized/sicoob/investimentos_sicoob_sample.pdf",
            "expected/sicoob_investments_expected.json",
        ),
    ),
)
def test_parser_matches_expected_golden_file(parser, fixture: str, expected: str) -> None:
    document = parser.parse(
        fixture_path(fixture),
        import_batch_id=IMPORT_BATCH_ID,
        source_file_id=SOURCE_FILE_ID,
    )

    assert document.to_expected_payload() == expected_payload(expected)
    assert document.import_batch_id == IMPORT_BATCH_ID
    assert document.source_file_id == SOURCE_FILE_ID
    assert document.records

    for record in document.records:
        assert record.source_type == document.source_type
        assert record.import_batch_id == IMPORT_BATCH_ID
        assert record.source_file_id == SOURCE_FILE_ID
        assert record.confidence_score > 0
        assert record.needs_review is False
        assert record.raw_reference


def test_parser_failure_is_controlled(tmp_path: Path) -> None:
    invalid_file = tmp_path / "account_statement_sample.csv"
    invalid_file.write_text("not;the;expected;format\n", encoding="utf-8")

    parser = MercadoLivreAccountStatementCsvParser()

    with pytest.raises(ParserError) as exc:
        parser.parse(invalid_file, import_batch_id=IMPORT_BATCH_ID)

    assert exc.value.code == "invalid_mercado_livre_csv"
    assert exc.value.to_dict()["message"]


def test_parser_error_masks_sensitive_raw_reference() -> None:
    error = ParserError(
        "invalid_pdf",
        "Falha ao ler CPF 123.456.789-09",
        raw_reference={
            "raw_text": "CPF 123.456.789-09 Conta 12345-6 Rua Exemplo, 100",
        },
    )

    payload = error.to_dict()

    assert "123.456.789-09" not in str(payload)
    assert "12345-6" not in str(payload)
    assert "Rua Exemplo" not in str(payload)
    assert "***.***.***-**" in str(payload)
    assert "<mascarado>" in str(payload)
    assert "<endereco mascarado>" in str(payload)


def test_b3_monthly_parser_accepts_official_bad_dimension_xlsx(tmp_path: Path) -> None:
    workbook_path = tmp_path / "relatorio-consolidado-mensal-2026-maio.xlsx"
    _write_b3_official_shape_workbook(workbook_path)
    _force_xlsx_dimensions_to_a1(workbook_path)

    document = B3MonthlyConsolidatedXlsxParser().parse(
        workbook_path,
        import_batch_id=IMPORT_BATCH_ID,
        source_file_id=SOURCE_FILE_ID,
    )

    assert document.payload["reference_month"] == "2026-05"
    assert document.payload["positions_by_class"] == {
        "acao": Decimal("123.45"),
        "etf": Decimal("50.55"),
        "renda_fixa": Decimal("200.0"),
    }
    assert document.payload["income_received_total"] == Decimal("8.75")
    assert document.payload["trades_count"] == 1
    assert len(document.records) == 5


def _write_b3_official_shape_workbook(path: Path) -> None:
    workbook = Workbook()
    actions = workbook.active
    actions.title = "Posição - Ações"
    actions.append(
        [
            "Produto",
            "Instituição",
            "Conta",
            "Código de Negociação",
            "CNPJ da Empresa",
            "Código ISIN / Distribuição",
            "Tipo",
            "Escriturador",
            "Quantidade",
            "Quantidade Disponível",
            "Quantidade Indisponível",
            "Motivo",
            "Preço de Fechamento",
            "Valor Atualizado",
        ]
    )
    actions.append(["ACAO3 - EMPRESA TESTE", "B3", "", "ACAO3", "", "", "", "", 1, 1, 0, "", 123.45, 123.45])

    etf = workbook.create_sheet("Posição - ETF")
    etf.append(
        [
            "Produto",
            "Instituição",
            "Conta",
            "Código de Negociação",
            "CNPJ do Fundo",
            "Código ISIN / Distribuição",
            "Tipo",
            "Quantidade",
            "Quantidade Disponível",
            "Quantidade Indisponível",
            "Motivo",
            "Preço de Fechamento",
            "Valor Atualizado",
        ]
    )
    etf.append(["ETF11 - FUNDO TESTE", "B3", "", "ETF11", "", "", "", 1, 1, 0, "", 50.55, 50.55])

    fixed_income = workbook.create_sheet("Posição - Renda Fixa")
    fixed_income.append(
        [
            "Produto",
            "Instituição",
            "Emissor",
            "Código",
            "Indexador",
            "Tipo de regime",
            "Data de Emissão",
            "Vencimento",
            "Quantidade",
            "Quantidade Disponível",
            "Quantidade Indisponível",
            "Motivo",
            "Contraparte",
            "Preço Atualizado MTM",
            "Valor Atualizado MTM",
            "Preço Atualizado CURVA",
            "Valor Atualizado CURVA",
        ]
    )
    fixed_income.append(["CDB - BANCO TESTE", "B3", "BANCO TESTE", "", "CDI", "", "", "", 1, 1, 0, "", "", 199.0, 199.0, 200.0, 200.0])

    income = workbook.create_sheet("Proventos Recebidos")
    income.append(["Produto", "Pagamento", "Tipo de Evento", "Instituição", "Quantidade", "Preço unitário", "Valor líquido"])
    income.append(["ACAO3 - EMPRESA TESTE", "15/05/2026", "Dividendo", "B3", 1, 8.75, 8.75])

    trades = workbook.create_sheet("Negociações")
    trades.append(
        [
            "Código de Negociação",
            "Período (Inicial)",
            "Período (Final)",
            "Instituição",
            "Quantidade (Compra)",
            "Quantidade (Venda)",
            "Quantidade (Líquida)",
            "Preço Médio (Compra)",
            "Preço Médio (Venda)",
        ]
    )
    trades.append(["ACAO3", "01/05/2026", "31/05/2026", "B3", 10, 0, 10, 12.34, 0])
    workbook.save(path)


def _force_xlsx_dimensions_to_a1(path: Path) -> None:
    patched = path.with_suffix(".patched.xlsx")
    with ZipFile(path, "r") as source, ZipFile(patched, "w") as target:
        for item in source.infolist():
            payload = source.read(item.filename)
            if item.filename.startswith("xl/worksheets/sheet") and item.filename.endswith(".xml"):
                payload = re.sub(b'<dimension ref="[^"]+"', b'<dimension ref="A1"', payload, count=1)
            target.writestr(item, payload)
    patched.replace(path)
